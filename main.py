import os
import time
from tkinter import *
from tkinter import ttk
from functools import partial
import subprocess
from threading import Thread
from threading import Timer
import openpyxl
import psutil
import pymysql
import datetime
from datetime import date
from multiprocessing import Process
from tkinter.messagebox import showinfo,showerror,showwarning
import multiprocessing
from multiprocessing import Pool

#список элементов для combobox-ов
modelForComboBox = []
nowCpu = 0

OZY= 0
MAX= 0
NOW = 0
qury = 0
MAXP = 0
inp = ""
NameBD = ""

class EditXLSS(Tk):
    def __init__(self):
        super().__init__()
        # тут я его настраивать начинаю
        self.title("Настройка запросов")
        self.geometry("920x500")

        self.CommandText = Text(width=90,height=20,font=("Arial Bold", 12))
        self.CommandText.grid(column=0, columnspan=3,row=1,pady=20,padx=35)

        self.btn1 = ttk.Button(self, text="Обновить данные")
        self.btn1["command"] = self.update
        self.btn1.grid(column=0,row=0,pady=20,padx=35)

        self.btn2 = ttk.Button(self, text="Сохранить данные")
        self.btn2["command"] = self.load
        self.btn2.grid(column=1, row=0, pady=20, padx=35)

        self.btn3 = ttk.Button(self, text="Назад в меню")
        self.btn3["command"] = self.back
        self.btn3.grid(column=2, row=0, pady=20, padx=35)

    def back(self):
        self.destroy()
        window = MainWindow()

    def load(self):
        test = self.CommandText.get('1.0', END)
        mas = []
        mas = test.split('\n')
        wb = openpyxl.load_workbook('sql.xlsx')
        sheet = wb.active

        for i in range(len(mas)):
            sheet["A" + str(i + 1)] = mas[i]
        wb.save('sql.xlsx')

    def update(self):
        gg = readxlsx()
        self.CommandText.delete('1.0', END)
        for i in range(len(gg)):
            self.CommandText.insert(INSERT, gg[i] + '\n')

#окно root
class WindowRoot(Tk):


    #метод для создания окна
    def __init__(self):
        super().__init__()

        #тут я его настраивать начинаю
        self.title("Администратор")
        self.geometry("1550x400")
        self.resizable(width=False, height=False)

        #Определяю столбцы для таблицы (нужно бужет это для каждой отдельной таблицы..
        columns = ("ID","name", "age", "email")

        #создаю сам элемент и вывожу его на экран через pack
        self.tree = ttk.Treeview(columns=columns, show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew",rowspan=3)

        #Переименовываю зоголовки т.е это нужно так-же для каждой отдельной таблицы
        self.tree.heading("ID",text="Первичный ключ")
        self.tree.heading("name", text="Имя")
        self.tree.heading("age", text="Возраст")
        self.tree.heading("email", text="Email")



        #кнопка для обновления данных, хотя можно и без неё
        btnAddDate = ttk.Button(self, text="Обновить данные")
        btnAddDate.grid(column=1, row=7, padx=15, pady=10)
        btnAddDate["command"] = self.refreshDate

        # кнопка для обновления данных, хотя можно и без неё
        btnDelDate = ttk.Button(self, text="Удалить данные")
        btnDelDate.grid(column=2, row=7, padx=15, pady=10)
        btnDelDate["command"] = self.clearTree

        #ещё кнопки в окне
        btnAdd = ttk.Button(self, text="Добавить")
        btnAdd.grid(column=3, row=0, padx=15, pady=10)
        btnAdd["command"] = self.addData

        btnDel = ttk.Button(self, text="Удалить")
        btnDel.grid(column=3, row=1, padx=15, pady=10)
        btnDel["command"] = self.delData

        btnEdit = ttk.Button(self, text="Изменить")
        btnEdit.grid(column=3, row=2, padx=15, pady=10)
        btnEdit["command"] = self.edtData

        btnBack = ttk.Button(self, text="Назад")
        btnBack.grid(column=3, row=3, padx=15, pady=10)
        btnBack["command"] = self.goBack

        self.laberOverComboBox = ttk.Label(self, text="Выберете таблицу",
                                     font=("Arial Bold", 12))
        self.laberOverComboBox.grid(column=3, row=4, pady=10, padx=50)
        # выпадающий список
        self.combobox = ttk.Combobox(values=modelForComboBox, width=20)
        self.combobox.grid(column=3, row=5, pady=10)
        # через partial передаю нормально аргументы в метод иначе оч плохо всё будет
        self.combobox.bind("<<ComboboxSelected>>", partial(self.selected))

        #поля ввода

        # //////////// ПЕРВЫЙ СТОЛБЕЦ

        self.entry1 = ttk.Entry(self, width=20, font=("Arial Bold", 12),show="some body")  # поле ввода
        self.entry1.grid(column=1, row=1,pady=10)


        #для отображения того, что будет вводить
        self.entryLabel1 = ttk.Label(self, text="Введите Имя",
                                            font=("Arial Bold", 12))
        self.entryLabel1.grid(column=1, row=0,pady=10, padx=50)

        #////////////

        self.entry2 = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.entry2.grid(column=1, row=3, pady=10)

        # для отображения того, что будет вводить
        self.entryLabel2 = ttk.Label(self, text="Введите Фамилию",
                                            font=("Arial Bold", 12))
        self.entryLabel2.grid(column=1, row=2, pady=10, padx=50)

        # ////////////

        self.entry3 = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.entry3.grid(column=1, row=5, pady=10)

        # для отображения того, что будет вводить
        self.entryLabel3 = ttk.Label(self, text="Введите Логин",
                                     font=("Arial Bold", 12))
        self.entryLabel3.grid(column=1, row=4, pady=10, padx=50)

        # //////////// ВТОРОЙ СТОЛБЕЦ

        self.entry4 = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.entry4.grid(column=2, row=1, pady=10)

        self.entryLabel4 = ttk.Label(self, text="Введите пароль",
                                     font=("Arial Bold", 12))
        self.entryLabel4.grid(column=2, row=0, pady=10, padx=50)

        # ////////////

        self.entry5 = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.entry5.grid(column=2, row=3, pady=10)

        # для отображения того, что будет вводить
        self.entryLabel5 = ttk.Label(self, text="Введите уровень доступа",
                                     font=("Arial Bold", 12))
        self.entryLabel5.grid(column=2, row=2, pady=10, padx=50)



        #это типо плейсхолдера т.к обычного нет в данной библиотеке

        # метод для выбора элемента из комбо бокса

    def selected(self, event):
        selection = self.combobox.get()
        print(selection)

    #метод для очистки данных в таблице
    def clearTree(self):
        print("Очищаю данные")
        x = self.tree.get_children()
        for item in x:  ## Changing all children from root item
            self.tree.delete(item)


    #метод для обновления данных
    def refreshDate(self):
        #добавляю данные из массива в таблицу\
        self.clearTree()
        print("Обновляю данные")

        con = pymysql.connect(host='localhost', user='adm', password='zekindaplaneta3900AS!',
                              database='practic_last_per_2')  # конект к бд
        cur = con.cursor()
        cur.execute("SELECT * FROM users")
        rows = cur.fetchall()
        for row in rows:
            self.tree.insert("", END, values=(row[0],row[1], row[2], row[3],row[8]))
        con.close()
        # for person in self.people:
        #     self.tree.insert("", END, values=person)

    #методы кнопок добавления удаления и изменения
    def addData(self):
        print("Добавляю данные")
        con = pymysql.connect(host='localhost', user='adm', password='zekindaplaneta3900AS!',
                              database='practic_last_per_2')  # конект к бд


        with con:
            with con.cursor() as cursor:
                try:
                    ima = self.entry1.get()
                    fam = self.entry2.get()
                    pas = self.entry4.get()
                    log = self.entry3.get()
                    dostup = self.entry5.get()
                    query = f"insert into practic_last_per_2.users(LONGNAME,FULLNAME,PASSW,SIGNNAME,ACCESS_LEVEL) values ('{ima}','{fam}','{pas}','{log}','{dostup}');"
                    cursor.execute(query)
                    con.commit()
                except Exception as e:
                    showerror("Ошибка от базы данных",str(e))

        self.refreshDate()

    def delData(self):
        print("Удаляю данные")
        selected_item = self.tree.selection()[0]
        selected_item = selected_item[3:]## get selected item
        con = pymysql.connect(host='localhost', user='adm', password='zekindaplaneta3900AS!',
                              database='practic_last_per_2')  # конект к бд

        with con:
            with con.cursor() as cursor:
                try:

                    query = f"delete from users where USER_ID = '{selected_item}';"
                    cursor.execute(query)
                    con.commit()
                except Exception as e:
                    showerror("Ошибка от базы данных", str(e))

        self.refreshDate()

    def edtData(self):
        print("Изменяю данные")
        selected_item = self.tree.selection()[0]
        selected_item = selected_item[3:]  ## get selected item
        con = pymysql.connect(host='localhost', user='adm', password='zekindaplaneta3900AS!',
                              database='practic_last_per_2')  # конект к бд

        with con:
            with con.cursor() as cursor:
                try:
                    ima = self.entry1.get()
                    fam = self.entry2.get()
                    pas = self.entry4.get()
                    log = self.entry3.get()
                    dostup = self.entry5.get()
                    query = f"update users set LONGNAME = '{ima}', FULLNAME = '{fam}',PASSW = '{pas}',ACCESS_LEVEL = '{dostup}',SIGNNAME = '{log}' WHERE USER_ID = '{selected_item}';"
                    cursor.execute(query)
                    con.commit()
                except Exception as e:
                    showerror("Ошибка от базы данных", str(e))

        self.refreshDate()


    #метод для кнопки, а именно переход обратно в главное окно
    def goBack(self):
        self.destroy()
        window = MainWindow()


class Avtorization(Tk):
    # метод для создания окна
    def __init__(self):
        super().__init__()

        # конфигурация окна
        self.title("Авторизация")
        self.geometry("250x370")
        self.resizable(width=False,height=False)

        #кнопки для переходв
        self.vvediteLogin = ttk.Label(self, text="Введите логин",  font=("Arial Bold", 12))
        self.vvediteLogin.grid(column=0,row=0,pady=20,padx=35)

        self.login = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.login.grid(column=0, row=1,pady=20,padx=35)

        self.vveditePassword = ttk.Label(self, text="Введите пароль", font=("Arial Bold", 12))
        self.vveditePassword.grid(column=0, row=2,pady=20,padx=35)

        self.password = ttk.Entry(self, width=20, font=("Arial Bold", 12), show="some body")  # поле ввода
        self.password.grid(column=0, row=3,pady=20,padx=35)

        self.buttonRoot = ttk.Button(self, text="Вход")
        self.buttonRoot["command"] = self.goToRootWindow
        self.buttonRoot.grid(column=0, row=4,pady=20,padx=35)

        self.buttonRootBack = ttk.Button(self, text="Назад")
        self.buttonRootBack["command"] = self.goToMenu
        self.buttonRootBack.grid(column=0, row=5,pady=20,padx=35)

    #методы переход между окнами

    def goToMenu(self):
        self.destroy()
        window = MainWindow()

    def goToRootWindow(self):
        adminLog = self.login.get()
        admingpas = self.password.get()
        if adminLog == "Admin" and admingpas == "Admin":
                self.destroy()
                window = WindowRoot()
        else:
            showwarning("Не корректные данные","Вы ввели не верный логин или пароль")

#главное окно
class MainWindow(Tk):

    # метод для создания окна
    def __init__(self):
        super().__init__()

        # конфигурация окна
        self.title("Меню")
        self.geometry("270x250")
        self.resizable(width=False,height=False)

        #кнопки для переходв
        self.button = ttk.Button(self, text="Переход в окно нагрузки ПК")
        self.button["command"] = self.goToCPUWindow
        self.button.grid(column=0,row=0,pady=20,padx=35)

        self.button1 = ttk.Button(self, text="Переход в окно нагрузки Сервера")
        self.button1["command"] = self.goToGPUWindow
        self.button1.grid(column=0,row=1,pady=20,padx=35)

        self.buttonRoot = ttk.Button(self, text="Переход в окно администратора")
        self.buttonRoot["command"] = self.goToRootWindow
        self.buttonRoot.grid(column=0, row=2,pady=20,padx=35)

        self.buttonRoot = ttk.Button(self, text="Переход в настройки файла запросов")
        self.buttonRoot["command"] = self.goToXLSS
        self.buttonRoot.grid(column=0, row=3, pady=20, padx=35)

    #методы переход между окнами

    def goToXLSS(self):
        self.destroy()
        window = EditXLSS()

    def goToCPUWindow(self):
        self.destroy()
        window = WindowCPU()

    def goToGPUWindow(self):
        self.destroy()
        window = WindowGPU()

    def goToRootWindow(self):
        self.destroy()
        window = Avtorization()







#окно нагрузки ЦПУ
class WindowCPU(Tk):

    #метод создания окна
    def __init__(self):
        super().__init__()


        #главная настройка окна
        self.title("Тест ПК")
        self.geometry("600x320")
        self.resizable(width=False, height=False)

        #элементы
        self.labelInfo1 = ttk.Label(self, text="Порог нагрузки ПК",
                           font=("Arial Bold", 12))
        self.labelInfo1.grid(column=0, row=0, pady=20, padx=10)

        #Ввод необходимой  нагрузки
        self.procents = ttk.Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода
        self.procents.grid(column=0, row=1, padx=20, pady=10)

        self.labelInfoUpper = ttk.Label(self, text="Текущая нагрузка\nв процентах",
                                   font=("Arial Bold", 12), justify="center")
        self.labelInfoUpper.grid(column=1, row=0, pady=20, padx=10)

        #Вывод текущего процента нагрузки ЦП
        self.labelInfo = ttk.Label(self, text="0",
                                   font=("Arial Bold", 12))
        self.labelInfo.grid(column=1, row=1, pady=20, padx=10)

        #вывод статуса программы
        self.labelStatus = ttk.Label(self, text="Остановлен",
                                   font=("Arial Bold", 12))
        self.labelStatus.grid(column=2, row=0)

        #не до конца понимаю как работает self, но если запихнуть его в кнопки, то они перестают работать
        btnStart = ttk.Button(self, text="Старт")
        btnStart.grid(column=0, row=3, padx=15, pady=10)
        btnStart["command"] = self.btnStart

        btnapdpk = ttk.Button(self, text="Обновление")
        btnapdpk.grid(column=1, row=3, padx=15, pady=10)
        btnapdpk["command"] = self.btnapdpk

        btnStop = ttk.Button(self, text="Остановить")
        btnStop.grid(column=2, row=3, padx=15, pady=10)
        btnStop["command"] = self.btnStop

        btnBack = ttk.Button(self, text="Назад")
        btnBack.grid(column=0, row=6, padx=15, pady=10)
        btnBack["command"] = self.btnBack


        self.labelPotokiMax = ttk.Label(self, text="MAX потоков",
                                     font=("Arial Bold", 12))
        self.labelPotokiMax.grid(column=3, row=0)


        #Вывод максимального колличества потоков
        self.labelPotokiMaxINFO = ttk.Label(self, text="1",
                                        font=("Arial Bold", 12))
        self.labelPotokiMaxINFO.grid(column=3, row=1)

        self.labelPotokiNOW = ttk.Label(self, text="NOW потоков",
                                        font=("Arial Bold", 12))
        self.labelPotokiNOW.grid(column=3, row=2)

        # Вывод текущих потоков
        self.labelPotokiNOWINFO = ttk.Label(self, text="1",
                                            font=("Arial Bold", 12))
        self.labelPotokiNOWINFO.grid(column=3, row=3)

        self.Requests = ttk.Label(self, text="Отправленных\nнагрузок",
                                        font=("Arial Bold", 12), justify="center")
        self.Requests.grid(column=3, row=4)

        # Вывод текущего колличества запросов
        self.RequestsShow = ttk.Label(self, text="1",
                                            font=("Arial Bold", 12))
        self.RequestsShow.grid(column=3, row=5, rowspan=2)


    #переход в главное окно
    def btnBack(self):
        self.destroy()
        window = MainWindow()

    def btnapdpk(self):



        f = open('currentProc.txt', 'r')
        lines = f.readlines()

        massive = []

        for line in lines:
            massive.append(line)

        f.close()

        thisMAX = massive[0]
        thisNOW = massive[1]
        thisqury = massive[2]

        thisMAX = thisMAX.replace('\n', '')
        thisNOW = thisNOW.replace('\n', '')

        print(MAX, "ITS BUTTON")
        self.labelInfo.config(text=str(GetCpuPersents()))
        # Вывод максимального колличества потоков
        self.labelPotokiMaxINFO.config(text=thisMAX)
        # Текущие потоки
        self.labelPotokiNOWINFO.config(text=thisNOW)
        # Вывод текущего колличества запросов
        self.RequestsShow.config(text=thisqury)

    #старт теста
    def btnStart(self):
        self.labelStatus.config(text="Старт")
        countOfCpu = multiprocessing.cpu_count()
        if (int(self.procents.get()) <= countOfCpu):
            self.labelStatus.config(text="Работает")
            with open('currentProc.txt', 'w') as file:
                print("Создание файла")
            global p
            p = proc_startcpu(self.procents.get())
        else:
            showerror(title="Ошибка",
                      message=f"Извините, но вам доступно только {countOfCpu - 1} процессоров для нагрузку\nА вы ввели {self.procents.get()}")


    #стоп
    def btnStop(self):
        self.labelStatus.config(text="Остановлен")
        os.remove('currentProc.txt')
        fullString = str(GetCpuPersents())
        self.labelInfo.config(text=fullString)
        self.labelPotokiMaxINFO.config(text="0")
        # Текущие потоки
        self.labelPotokiNOWINFO.config(text="0")
        # Вывод текущего колличества запросов
        self.RequestsShow.config(text="0")
        proc_stopcpu(p)

def GetCpuPersents():
        output = str(subprocess.check_output('wmic cpu get loadpercentage'))
        nowCpu = int(output[24] + output[25])
        return nowCpu

def threadscpu(maxLimit):
    intMaxLimit = int(maxLimit)
    i = 0
    T = intMaxLimit  # стартовое число ядер
    countermax = T

    sended_request = 0
    proc = Timer(1, GetCpuPersents, args=())
    proc.start()
    while True:
        #Загружает T колличество процессоров до придела нагружая его..
        pool = Pool(T)
        pool.map(MakePCWork,range(T))
        sended_request += T
        time = Timer(0.1, log, args=(T, countermax, sended_request,))
        time.start()
        time = Timer(0.1, setLogsInTXT, args=(T, countermax, sended_request,))
        time.start()



def MakePCWork(workX):
    x = workX * workX

def proc_startcpu(inpt):
    proc_startcpu = Process(target=threadscpu,args=(int(inpt),),daemon=False)
    proc_startcpu.start()
    return proc_startcpu


def proc_stopcpu(proc_stopcpu):
    proc_stopcpu.kill()


selection = "select * from users"

#окно нагрузки GPU
class WindowGPU(Tk):
    # метод создания окна
    def __init__(self):
        super().__init__()

        modelForComboBox = readxlsx()


        # главная настройка окна
        self.title("Тест Сервера")
        self.geometry("600x320")
        self.resizable(width=False, height=False)

        # элементы
        self.labelInfo = ttk.Label(self, text="Порог нагрузки Сервера",
                                   font=("Arial Bold", 12))
        self.labelInfo.grid(column=0, row=0, pady=20, padx=10)

        self.procents = ttk.Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода
        self.procents.insert(0, "")
        self.procents.grid(column=0, row=1, padx=20, pady=10)

        self.labelInfoUpper = ttk.Label(self, text="Текущая нагрузка\nв процентах",
                                   font=("Arial Bold", 12), justify="center")
        self.labelInfoUpper.grid(column=1, row=0, pady=20, padx=10)

        self.labelInfo = ttk.Label(self, text="60",
                                   font=("Arial Bold", 12))
        self.labelInfo.grid(column=1, row=1, pady=20, padx=10)

        self.labelStatus = ttk.Label(self, text="Остановлен",
                                     font=("Arial Bold", 12))
        self.labelStatus.grid(column=2, row=0, rowspan=2)

        # не до конца понимаю как работает self, но если запихнуть его в кнопки, то они перестают работать
        btnStart = ttk.Button(self, text="Старт")
        btnStart.grid(column=0, row=3, padx=15, pady=10)
        btnStart["command"] = self.btnStart

        btnapd = ttk.Button(self, text="Обновление")
        btnapd.grid(column=1, row=3, padx=15, pady=10)
        btnapd["command"] = self.btnapd


        btnStop = ttk.Button(self, text="Остановить")
        btnStop.grid(column=2, row=3, padx=15, pady=10)
        btnStop["command"] = self.btnStop

        btnBack = ttk.Button(self, text="Назад")
        btnBack.grid(column=0, row=6, padx=15, pady=10)
        btnBack["command"] = self.btnBack

        self.labelInfoAboutComboBox = ttk.Label(self, text="Назване базы данных",
                                                font=("Arial Bold", 12))
        self.labelInfoAboutComboBox.grid(column=0, row=4, columnspan=3, pady=10)

        #/////////////////////////////////////////
        #////////////////////////////////////////
        self.entryNameDb = ttk.Entry(self, width=20, font=("Arial Bold", 12))  # поле ввода
        self.entryNameDb.grid(column=0, row=5, columnspan=3, pady=10)
        # через partial передаю нормально аргументы в метод иначе оч плохо всё будет
        self.entryNameDb.insert(0,"")#СЮДА ВСТАВИТЬ ВМЕСТО КАВЫЧЕК СПИСОК С ДЕШИФРОВАННЫМИ ДАННЫМИ

        #В этом-же окне заменить комбобок на это

        #///////////////////////////////////
        #///////////////////

        self.labelPotokiMax = ttk.Label(self, text="MAX потоков",
                                        font=("Arial Bold", 12))
        self.labelPotokiMax.grid(column=3, row=0, rowspan=2)

        # Вывод максимального колличества потоков
        self.labelPotokiMaxINFO = ttk.Label(self, text="1",
                                            font=("Arial Bold", 12))
        self.labelPotokiMaxINFO.grid(column=3, row=1, rowspan=2)

        self.labelPotokiNOW = ttk.Label(self, text="NOW потоков",
                                        font=("Arial Bold", 12))
        self.labelPotokiNOW.grid(column=3, row=2, rowspan=2)

        # Вывод текущих потоков
        self.labelPotokiNOWINFO = ttk.Label(self, text="1",
                                            font=("Arial Bold", 12))
        self.labelPotokiNOWINFO.grid(column=3, row=3, rowspan=2)

        self.Requests = ttk.Label(self, text="Отправленных\nзапросов",
                                  font=("Arial Bold", 12), justify="center")
        self.Requests.grid(column=3, row=4, rowspan=2)

        # Вывод текущего колличества запросов
        self.RequestsShow = ttk.Label(self, text="1",
                                      font=("Arial Bold", 12))
        self.RequestsShow.grid(column=3, row=5, rowspan=2)


    # метод для выбора элемента из комбо бокса
    def selected(self, event):
        global selection
        selection = self.combobox.get()
        print(selection)

    # переход в главное окно
    def btnBack(self):
        self.destroy()
        window = MainWindow()

    def btnapd(self):
        f = open('currentProc.txt', 'r')
        lines = f.readlines()
        massive = []

        for line in lines:
            massive.append(line)
        f.close()
        thisMAX = massive[0]
        thisNOW =  massive[1]
        thisqury =  massive[2]
        thisMAX = thisMAX.replace('\n','')
        thisNOW = thisNOW.replace('\n', '')
        print(MAX, "ITS BUTTON")
        self.labelInfo.config(text=str(psutil.virtual_memory()[2]))
        # Вывод максимального колличества потоков
        self.labelPotokiMaxINFO.config(text=thisMAX)
        # Текущие потоки
        self.labelPotokiNOWINFO.config(text=thisNOW)
        # Вывод текущего колличества запросов
        self.RequestsShow.config(text=thisqury)

    # старт теста
    def btnStart(self):
        global selection
        if(selection != ""):
            with open('currentProc.txt', 'w') as file:
                print("Создание файла")
            self.labelStatus.config(text="Старт")
            global inp
            inp = self.procents.get()

            global p
            p = proc_start(self.entryNameDb.get())
            #ProgramDelay()

        else:
            showwarning(title="Предупреждение",
                      message=f"Извините, но вы не выбрали какой запрос будут отправлять в базу данных")


    # стоп
    def btnStop(self):
        self.labelStatus.config(text="Остановлен")
        ProgramDelay()
        proc_stop(p)
        os.remove('currentProc.txt')
        self.labelInfo.config(text=str(psutil.virtual_memory()[2]))
        self.labelPotokiMaxINFO.config(text="0")
        # Текущие потоки
        self.labelPotokiNOWINFO.config(text="0")
        # Вывод текущего колличества запросов
        self.RequestsShow.config(text="0")


def proc_start(BD):
    global inp
    p_to_start = Process(target=threads,args=(int(inp),BD,),daemon=False)
    p_to_start.start()
    return p_to_start


def proc_stop(p_to_stop):
    p_to_stop.kill()


def setLogsInTXT(sended_request,num,nummax):
        f = open('currentProc.txt', 'r+')
        f.write(f"{sended_request}\n{num}\n{nummax}")
        f.close()


def threads(porog, BD):
    xlsx = readxlsx()
    T = 5  # стартовое число потоков
    countermax = T
    sended_request = 0
    while True:
        time = Timer(0.1, setLogsInTXT, args=(T, countermax, sended_request,))
        time.start()
        if int(psutil.virtual_memory()[2]) > porog:  # верхний порог нагрузки
            T -= 1
        if int(psutil.virtual_memory()[2]) < porog:
            T += 1
        if T > countermax:  #
            countermax = T
        if T <= 0:
            T = 1
        time = Timer(5, log, args=(T, countermax, sended_request,))
        time.start()
        threads = []
        for n in range(int(T)):
            t = Thread(target=bd, args=(xlsx,BD,), daemon=False)
            t.start()
            threads.append(t)
        for t in threads:
            t.join()
            sended_request = sended_request + 1
            OZY = str(psutil.virtual_memory()[2])
            MAX = countermax
            NOW = T




def log(num,nummax,sended_request):
    current_date = date.today()
    dt_now = datetime.datetime.now()
    f = open(str(current_date)+'.txt','a')
    f.write(str(dt_now) + " active_thread: " + str(num)  + " max_thread: " + str(nummax) + " sended_request: " + str(sended_request) + "\n")
    f.close()

def logForCpu(num,nummax,sended_request):
    current_date = date.today()
    dt_now = datetime.datetime.now()
    f = open(str(current_date)+'.txt','a')
    f.write(str(dt_now) + " active_thread: " + str(num)  + " max_thread: " + str(nummax) + " sended_request: " + str(sended_request) + "\n")
    f.close()
    SaveToDateBaseData(num,0,sended_request)


def readxlsx():
    wb = openpyxl.load_workbook('sql.xlsx')
    sheet = wb.active
    xlsx = []
    for i in range(sheet.max_row):
        xlsx.append(str(sheet["A" + str(i + 1)].value))
    return xlsx

def bd(xls, dnName):
    user = "test" #decryptedAccses[1]
    password = "test"#decryptedAccses[0]
    connection = cx_Oracle.connect(user + "/" + password + "@" + dnName, encoding="UTF-8")
    print(connection)
    cursor = connection.cursor()

    for i in xls:
        dnss = f"""
                   {i}"""
        cursor.execute(dnss)
        ver = cursor.fetchone()



def SaveToDateBaseData(varnum,start_num,end_num):
    con = pymysql.connect(host='localhost', user='adm', password='zekindaplaneta3900AS!',
                          database='practic_last_per_2')  # конект к бд
    with con:
        with con.cursor() as cursor:
            query = f"INSERT INTO doc_nums (VARNUMM, DOC_USER_ID,START_NUM,END_NUM) VALUES ('{varnum}', null, '{start_num}', '{end_num}');"

            cursor.execute(query)
            con.commit()




def ProgramDelay():
    global p
    from datetime import datetime
    now = datetime.now()
    from datetime import timedelta
    run_at = now + timedelta(seconds=10)
    delay = (run_at - now).total_seconds()
    proc_stopcpu(p)
    time.sleep(delay)

    p = proc_start("BD")






#это код для запуска приложения, так сказать главное окно для начала переходов (костыли)
def click():
    root.destroy()
    window = MainWindow()

if __name__ == '__main__':
    root = Tk()
    root.title("Окно запуска")
    root.geometry("250x200")
    open_button = ttk.Button(text="Запуск", command=click)
    open_button.pack(anchor="center", expand=1)
    root.mainloop()